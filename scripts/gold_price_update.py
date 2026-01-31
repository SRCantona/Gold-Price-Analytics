import os
import re
import requests
import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook

# =========================
# CONFIG
# =========================
EXCEL_PATH = r"C:/Users/saleh/Desktop/data/gold_prices.xlsx"
# على GitHub Actions استخدم:
# EXCEL_PATH = "data/gold_prices.xlsx"

SOURCE_URL = "https://saudigoldprice.com/"

KARAT_COLS = ["SAR_24K", "SAR_22K", "SAR_21K", "SAR_18K"]

# ✅ أول تشغيل يبدأ من هذا التاريخ فقط
BOOTSTRAP_START_DATE = "2026-01-23"


# =========================
# EXCEL HELPERS
# =========================
def load_daily_any_sheet(path: str) -> pd.DataFrame:
    """Load Daily sheet if exists; otherwise return empty DataFrame with standard columns."""
    if not os.path.exists(path):
        return pd.DataFrame(columns=["Date"] + KARAT_COLS)

    wb = load_workbook(path)
    sheets = wb.sheetnames
    sheet = "Daily_Data" if "Daily_Data" in sheets else sheets[0]

    df = pd.read_excel(path, sheet_name=sheet)

    if "Date" not in df.columns:
        return pd.DataFrame(columns=["Date"] + KARAT_COLS)

    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"])

    for c in KARAT_COLS:
        if c not in df.columns:
            df[c] = pd.NA

    return df[["Date"] + KARAT_COLS].copy()


def get_start_date(daily: pd.DataFrame) -> str:
    """
    - إذا الإكسل فاضي: يبدأ من 2026-01-23
    - إذا فيه بيانات: يبدأ من آخر يوم + 1
    """
    if daily is None or len(daily) == 0:
        return BOOTSTRAP_START_DATE

    last_date = daily["Date"].max()
    return (last_date + timedelta(days=1)).strftime("%Y-%m-%d")


def write_yearly_note(excel_path: str):
    """
    Adds a note at the top of the Yearly_Averages sheet:
    Yearly values are YTD (year-to-date) average based on available days so far.
    """
    wb = load_workbook(excel_path)
    if "Yearly_Averages" not in wb.sheetnames:
        wb.save(excel_path)
        return

    ws = wb["Yearly_Averages"]

    # avoid duplicating note if rerun
    existing = ws["A1"].value
    if existing and "Year-to-date" in str(existing):
        wb.save(excel_path)
        return

    ws.insert_rows(1)
    ws["A1"] = (
        "NOTE: Yearly_Averages = Year-to-date (YTD) average based on available days so far in the year "
        "(not a full completed year)."
    )

    max_col = ws.max_column
    if max_col >= 2:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

    wb.save(excel_path)


# =========================
# SCRAPER (TODAY PRICES)
# =========================
def fetch_today_from_saudigoldprice() -> pd.DataFrame:
    """
    Fetch today's gold prices (SAR per gram) for 24/22/21/18 from saudigoldprice.com.
    Fixes encoding issues and uses robust regex patterns.
    """
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/120.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "ar,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }

    r = requests.get(SOURCE_URL, headers=headers, timeout=30)
    r.raise_for_status()

    # ✅ Fix encoding (prevents Ø£Ø³Ø¹Ø§Ø±...)
    if not r.encoding:
        r.encoding = r.apparent_encoding or "utf-8"
    # كثير مواقع عربية ترجع بدون encoding صحيح — نجبره:
    r.encoding = "utf-8"

    html = r.text

    def _clean_number(s: str) -> float:
        s = s.replace(",", "").strip()
        s = re.sub(r"[^0-9\.]", "", s)
        return float(s)

    def find_price(karat: int):
        """
        Search for price near Arabic labels like:
        'عيار 24' or 'جرام الذهب عيار 24'
        We capture the first number appearing shortly after the label.
        """
        patterns = [
            rf"جرام\s+الذهب\s+عيار\s*{karat}[^0-9]{{0,120}}([0-9\.,]+)",
            rf"سعر\s+جرام\s+الذهب\s+عيار\s*{karat}[^0-9]{{0,120}}([0-9\.,]+)",
            rf"عيار\s*{karat}[^0-9]{{0,120}}([0-9\.,]+)",
            rf"{karat}\s*قيراط[^0-9]{{0,120}}([0-9\.,]+)",
            rf"{karat}\s*K[^0-9]{{0,120}}([0-9\.,]+)",
        ]

        for pat in patterns:
            m = re.search(pat, html, flags=re.IGNORECASE | re.DOTALL)
            if m:
                try:
                    return _clean_number(m.group(1))
                except:
                    continue
        return None

    p24 = find_price(24)
    p22 = find_price(22)
    p21 = find_price(21)
    p18 = find_price(18)

    today = pd.to_datetime(datetime.now().date())

    if None in (p24, p22, p21, p18):
        # Diagnose: show small readable snippet
        snippet = re.sub(r"\s+", " ", html)[:800]
        raise RuntimeError(
            "Couldn't parse prices from saudigoldprice.com.\n"
            "Possible reasons:\n"
            "- Site markup changed\n"
            "- Bot protection served a different page\n"
            "- Prices are loaded via JavaScript in your region\n"
            f"\nHTML snippet:\n{snippet}"
        )

    return pd.DataFrame([{
        "Date": today,
        "SAR_24K": p24,
        "SAR_22K": p22,
        "SAR_21K": p21,
        "SAR_18K": p18,
    }])


# =========================
# MAIN
# =========================
def update_excel_sar_prices():
    daily = load_daily_any_sheet(EXCEL_PATH)

    start_date = get_start_date(daily)
    start_dt = pd.to_datetime(start_date)
    today_dt = pd.to_datetime(datetime.now().date())

    print(f"Updating from {start_dt.date()} to {today_dt.date()}")

    # لو اليوم قبل start_date، يعني ما فيه تحديث
    if today_dt < start_dt:
        print("No new data needed.")
        return

    # نجلب سعر اليوم فقط
    new_rows = fetch_today_from_saudigoldprice()

    # فلترة: إذا سعر اليوم أصلاً قبل start_date ما نضيفه
    new_rows = new_rows[new_rows["Date"] >= start_dt]
    if new_rows.empty:
        print("No new rows to add.")
        return

    # دمج وتوحيد
    daily = pd.concat([daily, new_rows], ignore_index=True, sort=False)
    daily = daily.drop_duplicates(subset=["Date"], keep="last").sort_values("Date")

    # Monthly averages
    monthly = (
        daily.set_index("Date")[KARAT_COLS]
        .resample("ME").mean().round(2).reset_index()
    )

    # Yearly averages = YTD average
    yearly = (
        daily.set_index("Date")[KARAT_COLS]
        .resample("YE").mean().round(2).reset_index()
    )

    os.makedirs(os.path.dirname(EXCEL_PATH) or ".", exist_ok=True)
    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="w") as writer:
        daily.to_excel(writer, sheet_name="Daily_Data", index=False)
        monthly.to_excel(writer, sheet_name="Monthly_Averages", index=False)
        yearly.to_excel(writer, sheet_name="Yearly_Averages", index=False)

    write_yearly_note(EXCEL_PATH)

    print("Excel updated successfully from saudigoldprice.com ✅")


if __name__ == "__main__":
    update_excel_sar_prices()
